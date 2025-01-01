import openpyxl as xl

wb = xl.load_workbook('/home/oseyemen/python/tutorial/automation_with_python/transactions.xlsx')
sheet = wb['Sheet1']


for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row, 3)
    print(cell.value)