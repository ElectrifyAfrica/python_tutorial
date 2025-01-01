import openpyxl as xl

wb = xl.load_workbook('/home/oseyemen/python/tutorial/automation_with_python/transactions.xlsx')
sheet = wb['Sheet1']


for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

    #print(cell.value)
    #print(corrected_price_cell.value)

wb.save('/home/oseyemen/python/tutorial/automation_with_python/transactions2.xlsx')